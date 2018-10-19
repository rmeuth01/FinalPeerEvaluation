import openpyxl



class Review:
    def __init__(self, name, score):
        self.name = name
        self.score = score
        self.count = 1

class Team:
    def __init__(self, team, name, score):
        self.teamnumber = team
        self.members = []
        self.updatemember(name, score)

    def teamtotal(self):
        total = 0
        totalcount = 0
        for member in self.members:
            total += member.score
            totalcount += member.count
        return total/totalcount

    def writescores(self, outfile):
        ideal = self.teamtotal()
        for member in self.members:
            scaling = (member.score/member.count) / ideal
            outfile.write(member.name+','+str(scaling)+'\n')

    def updatemember(self, name, score):
        for member in self.members:
            if member.name == name:
                member.score += score
                member.count += 1
                return
        self.members.append(Review(name, score))


    
    
class PeerAnalysis:
    def __init__(self, input_filename):
        self.teams = []
        self.input_filename = input_filename
        self.wb = openpyxl.load_workbook(self.input_filename)
        self.project_sheet = self.wb.get_sheet_by_name("Form Responses 1")
        self.current_row = 2 # 1-based indexing, skip the headers.
        while self.current_row <= self.project_sheet.max_row:
            #Data
            # 1 - Time Stamp
            # 2 - User Email
            # 3 - Team Number
            # 4 - Reviewee
            # 5 - Score 1
            # 6 - Score 2
            # 7 - Score 3
            # 8 - Score 4
            # 9 - Score 5
            # 10 - Score 6
            score = 0
            for col in range (5,11):
                score += int(self.project_sheet.cell(row=self.current_row, column=col).value)
           
            name = self.project_sheet.cell(row=self.current_row, column=4).value
            team = self.project_sheet.cell(row=self.current_row, column=3).value
            # Update Team
            self.updateteam(team, name, score)
            self.current_row += 1

    def writeout(self, output_filename):
        outfile = open(output_filename,'w')
        for team in self.teams:
            team.writescores(outfile)
        outfile.close()

    def updateteam(self, teamnumber, name, score):
        for team in self.teams:
            if team.teamnumber == teamnumber:
                #print('Found Team!')
                team.updatemember(name, score)
                return
        self.teams.append(Team(teamnumber, name, score))


# main
Analysis = PeerAnalysis("FSE100A Project Spyn Peer Evaluation SP18.xlsx")
Analysis.writeout("FSE100A_PeerScales.csv")

